# -*- coding: utf-8 -*-
"""补充匹配表缺失产品（2026-08-03）：从 26-7/26-8 数据源提取 150 个缺失产品的属性，
按现有匹配表格式规范生成补充行。先预览（--preview），确认后 --write 落盘。"""
import pandas as pd, re, sys, os

BASE = r'D:\E盘文件\电商渠道业绩看板'
TARGET = os.path.join(BASE, '各渠道月度目标数据.xlsx')
DATA = os.path.join(BASE, '各渠道销售数据源')
KEY = ['直销_伊稻_电商_天猫ITO旗舰店','直销_乐绘_电商_京东ITO京东自营旗舰店','直销_伊稻_电商_抖音ITO旗舰店','直销_乐绘_电商_抖音ITO官方旗舰店','直销_伊远_电商_抖音ITO行李箱旗舰店','直销_伊稻_电商_小红书ITO旗舰店','直销_伊稻_电商_视频号ITO旗舰店']

# 1. 读现有匹配表
m = pd.read_excel(TARGET, sheet_name='系列产品匹配')
match_names = set(str(x).strip() for x in m['货品名称'].dropna())
# 现有系列 -> (简称, 人群) 众数映射
series_info = {}
for _, r in m.iterrows():
    s = str(r.get('系列', '')).strip()
    if s and s != 'nan' and s not in series_info:
        short = str(r.get('简称-商品', '')).strip()
        aud = str(r.get('人群', '')).strip()
        series_info[s] = {'short': '' if short == 'nan' else short,
                          'aud': '' if aud == 'nan' else aud}

# 2. 收集缺失产品（数据源 26-8 优先）
rows = {}
for f in ['26-8', '26-7']:
    fp = os.path.join(DATA, f + '.xlsx')
    if not os.path.exists(fp):
        continue
    df = pd.read_excel(fp)
    df = df[df['店铺'].isin(KEY)]
    for _, r in df.iterrows():
        n = str(r.get('货品名称', '')).strip()
        if not n or n in match_names or n in rows:
            continue
        rows[n] = {'code': r.get('商家编码'), 'price': r.get('零售价')}

# 3. 系列级属性规则
SERIES_RULES = {
    'CLASSIC 15': {'cat': '行李箱', 'short': '15', 'aud': '', 'trunk': '非TRUNK箱型', 'prefix': ''},
    'CLASSIC NEAT': {'cat': '行李箱', 'short': 'NEAT', 'aud': '', 'trunk': 'TRUNK箱型', 'prefix': '压纹'},
    'CLASSIC WAVE': {'cat': '行李箱', 'short': 'WAVE', 'aud': '潮奢原生代', 'trunk': '非TRUNK箱型', 'prefix': '亮面'},
    'GINKGO WEEKEND': {'cat': '小型箱', 'short': 'WEEKEND', 'aud': '', 'trunk': '', 'prefix': '亮面'},
    'GINKGO WEEKEND BONPOINT': {'cat': '行李箱', 'short': 'WEEKEND-BP', 'aud': '云端商务家', 'trunk': '', 'prefix': '亮面'},
    'PISTACHIO 2 STRIPED': {'cat': '行李箱', 'short': 'P2', 'aud': '摩登新贵女', 'trunk': '非TRUNK箱型', 'prefix': '亮面'},
    'MYCENA TOTE': {'cat': '包袋', 'short': 'MYCENA', 'aud': '摩登新贵女', 'trunk': '', 'prefix': 'CORDURA'},
    'COSMETIC BAG': {'cat': '包袋', 'short': '', 'aud': '', 'trunk': '', 'prefix': 'CORDURA'},
    'PHONE POUCH': {'cat': '旅行箱配件', 'short': '', 'aud': '', 'trunk': '', 'prefix': 'CORDURA'},
    'MUSHROOM ORGANIZER 2': {'cat': '包袋', 'short': '', 'aud': '', 'trunk': '', 'prefix': '尼龙'},
    'LUGGAGE COVER': {'cat': '旅行箱配件', 'short': '', 'aud': '', 'trunk': '', 'prefix': ''},
    'LUGGAGE LOCK': {'cat': '旅行箱配件', 'short': '', 'aud': '', 'trunk': '', 'prefix': ''},
    'ORGANIZER': {'cat': '旅行箱配件', 'short': '', 'aud': '', 'trunk': '', 'prefix': ''},
    'PISTACHIO 箱套': {'cat': '旅行箱配件', 'short': '', 'aud': '', 'trunk': '', 'prefix': ''},
    'LUGGAGE TAG 2': {'cat': '行李牌', 'short': '', 'aud': '', 'trunk': '', 'prefix': '超纤'},
    'LUGGAGE TAG 2 ICON': {'cat': '行李牌', 'short': '', 'aud': '', 'trunk': '', 'prefix': '硅胶'},
}

def parse_name(name):
    """解析货品名称 -> (系列, 颜色, 尺寸, 品类, 箱型, 简称, 人群)"""
    mm = re.search(r'ITO\s+(.+?)系列', name)
    series_raw = mm.group(1).strip() if mm else ''
    # 归一化系列名
    series = series_raw
    # 特殊处理：BONPOINT
    is_bp = 'BONPOINT' in name
    if is_bp and series_raw.startswith('GINKGO WEEKEND'):
        series = 'GINKGO WEEKEND BONPOINT'
    # 系列规则
    rule = SERIES_RULES.get(series)
    if rule is None:
        # 尝试模糊匹配现有系列
        if series in series_info:
            rule = {'cat': '', 'short': series_info[series]['short'], 'aud': series_info[series]['aud'],
                    'trunk': '', 'prefix': ''}
        else:
            # 未知系列：从名称推断品类
            cat = '行李箱' if '旅行箱' in name and '配件' not in name else \
                  ('包袋' if ('包袋' in name or 'TOTE' in series or 'BAG' in series) else \
                   ('小型箱' if '18英寸' in name else \
                    ('行李牌' if ('TAG' in series or '行李牌' in name) else \
                     ('旅行箱配件' if '配件' in name else \
                      ('虚拟卡券' if '虚拟卡券' in series else \
                       ('包材辅料' if '包材' in series else \
                        ('维修配件' if ('维修' in name or '拉杆' in series or '锁' in series) else \
                         ('赠品' if '赠品' in name else '旅行生活'))))))))
            rule = {'cat': cat, 'short': '', 'aud': '', 'trunk': '', 'prefix': ''}
    # 取系列后的剩余文本
    rest = name[len('ITO ' + series_raw + '系列 '):] if '系列 ' in name else name
    # 颜色/尺寸解析
    color, size = '', ''
    cat = rule['cat']
    if cat == '行李箱' or cat == '小型箱':
        size_m = re.search(r'(\d+(?:\.\d+)?英寸)', rest)
        size = size_m.group(1) if size_m else ''
        # 颜色：去掉尺寸后的剩余中，去材质前缀（亮面/压纹/磨砂），取第一个颜色词
        pre = rest.split(size)[0] if size else rest
        for pfx in ['亮面', '压纹', '磨砂', '轻量', 'CORDURA', '尼龙', 'TPU半透明']:
            if pre.startswith(pfx):
                pre = pre[len(pfx):]
                break
        color = pre.strip()
        # TRUNK 特殊
        if 'TRUNK' in name:
            cat_ok = '行李箱'
        trunk = rule['trunk'] if rule['trunk'] else ('TRUNK箱型' if 'TRUNK' in name else '非TRUNK箱型')
    elif cat == '包袋':
        size_m = re.search(r'(\d+(?:\.\d+)?L)', rest)
        size = size_m.group(1) if size_m else ''
        pre = rest.split(size)[0] if size else rest
        for pfx in ['CORDURA', '尼龙', '涂层涤纶', '超纤']:
            if pre.startswith(pfx):
                pre = pre[len(pfx):]
                break
        color = pre.replace('旅行生活', '').strip()
        trunk = ''
    elif cat == '行李牌':
        # 超纤/硅胶前缀
        for pfx in ['超纤', '硅胶', '皮革']:
            if rest.startswith(pfx):
                rest = rest[len(pfx):]
                break
        if '旅行生活' in rest:
            pre = rest.split('旅行生活')[0].strip()
            color = pre
            size = ''
        elif '赠品' in rest:
            parts = rest.split('赠品')
            color = parts[0].strip()
            size = parts[1].strip() if len(parts) > 1 else ''
        # ICON 字母/图形
        icon_m = re.search(r'([A-Z]$|五角星|爱心)$', name)
        if icon_m:
            size = icon_m.group(1)
        trunk = ''
    elif cat == '旅行箱配件':
        size_m = re.search(r'(\d+(?:\.\d+)?(?:英寸|L))', rest)
        size = size_m.group(1) if size_m else ''
        pre = rest.split(size)[0] if size else rest
        # 颜色=材质或颜色描述
        color = pre.replace('旅行箱配件', '').replace('赠品', '').strip()
        if not color:
            color = ''
        trunk = ''
    else:
        # 虚拟卡券/包材辅料/维修配件/赠品/旅行生活
        size_m = re.search(r'((?:\d+(?:\.\d+)?(?:英寸|L|ML|天))|M号|S号|[A-Z]$|T20|001)', rest)
        size = size_m.group(1) if size_m else ''
        pre = rest.split(size)[0] if size else rest
        color = pre.replace('赠品', '').replace('维修配件', '').replace('包材辅料', '').strip()
        trunk = ''
    # 潮爸人群
    aud = rule['aud']
    if not aud and ('潮爸' in name or 'BONPOINT' in name):
        aud = '潮奢原生代' if '潮爸' in name else '云端商务家'
    if not color:
        color = ''
    return {'series': series, 'color': color, 'size': size, 'cat': cat,
            'trunk': trunk, 'short': rule['short'], 'aud': aud}

# 4. 生成补充行
new_rows = []
for name in sorted(rows.keys()):
    info = parse_name(name)
    new_rows.append({
        '商家编码': rows[name]['code'],
        '货品名称': name,
        '系列': info['series'],
        '颜色': info['color'],
        '尺寸': info['size'],
        '品类': info['cat'],
        '吊牌价': rows[name]['price'],
        '箱型': info['trunk'],
        '简称-商品': info['short'],
        '人群': info['aud'],
    })

df_new = pd.DataFrame(new_rows)
if '--preview' in sys.argv or len(sys.argv) == 1:
    print(f'共生成 {len(df_new)} 行，品类分布：')
    print(df_new['品类'].value_counts().to_string())
    print('\n=== 前30行预览 ===')
    for _, r in df_new.head(30).iterrows():
        print(f'  {r["货品名称"]} | 系列={r["系列"]} 颜色={r["颜色"]} 尺寸={r["尺寸"]} 品类={r["品类"]} 箱型={r["箱型"]} 简称={r["简称-商品"]} 人群={r["人群"]} | code={r["商家编码"]} price={r["吊牌价"]}')
    print(f'\n=== 颜色为空的行 ===')
    for _, r in df_new[df_new['颜色'] == ''].iterrows():
        print(f'  ⚠️ {r["货品名称"]} | 系列={r["系列"]} 尺寸={r["尺寸"]}')
if '--write' in sys.argv:
    import openpyxl
    wb = openpyxl.load_workbook(TARGET)
    ws = wb['系列产品匹配']
    start_row = ws.max_row + 1
    # 列：商家编码 货品名称 系列 颜色 尺寸 品类 吊牌价 (跳过8) 箱型 简称-商品 人群
    colmap = {'商家编码': 1, '货品名称': 2, '系列': 3, '颜色': 4, '尺寸': 5, '品类': 6,
              '吊牌价': 7, '箱型': 9, '简称-商品': 10, '人群': 11}
    n = 0
    for _, r in df_new.iterrows():
        for k, ci in colmap.items():
            v = r.get(k)
            if v == '' or (isinstance(v, float) and pd.isna(v)):
                v = None
            ws.cell(row=start_row + n, column=ci, value=v)
        n += 1
    wb.save(TARGET)
    print(f'\n✅ 已追加 {n} 行到「系列产品匹配」(行 {start_row}~{start_row+n-1})，匹配表 {start_row-1} -> {start_row+n-1} 行')
