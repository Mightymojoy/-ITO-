# -*- coding: utf-8 -*-
"""
从 各渠道销售数据源/YY-M.xlsx 同步实际销售额到 各渠道月度目标数据.xlsx 的 B列(实际销售)
使用流程：用户更新月度数据源后，运行此脚本同步到目标文件，再运行 刷新数据.bat
"""
import pandas as pd
import openpyxl
import os, time, sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import datetime, timedelta

BASE = r'D:\E盘文件\电商渠道业绩看板'
TARGET_FILE = os.path.join(BASE, '各渠道月度目标数据.xlsx')

# 只处理最近 N 天内有修改的文件（默认7天）
RECENT_DAYS = 7

# ===== 渠道映射：数据源中的店铺名 → 目标文件sheet名 =====
STORE_TO_SHEET = {
    '直销_伊稻_电商_天猫ITO旗舰店': '直销_伊稻_电商_天猫ITO旗舰店',
    '直销_乐绘_电商_京东ITO京东自营旗舰店': '直销_乐绘_电商_京东ITO京东自营旗舰店',
    '直销_伊稻_电商_抖音ITO旗舰店': ' 抖音ITO旗舰店（摩登新贵女）',
    '直销_乐绘_电商_抖音ITO官方旗舰店': ' 抖音ITO官方旗舰店（轻熟质享客）',
    '直销_伊远_电商_抖音ITO行李箱旗舰店': ' 抖音ito箱包旗舰店（云端商务家）',
    '直销_伊稻_电商_小红书ITO旗舰店': '小红书自营',
    '直销_伊稻_电商_视频号ITO旗舰店': '视频号',
}

# 需要排除的店铺（非看板渠道）
EXCLUDE_STORES = {
    '经销_伊稻_电商_天猫ITO奥莱旗舰店_上海韶瑶电子商务',
    '经销_伊稻_电商_京东ITO旗舰店_衍悠贸易（上海）',
    '经销_乐绘_电商_汕头市潮银贸易',
    '经销_乐绘_电商_同齐贸易（上海）_天猫ITO同齐专卖店',
    '联营_乐绘_北京和明兴_ITO武汉国际广场店',
    '经销_乐绘_电商_上海信橙贸易_ITO京东奥莱旗舰店',
    '联营_乐绘_北京和明兴_ITO北京国贸店',
    '经销_伊稻_航空商城_中国南航集团线下',
    '联营_乐绘_北京和明兴_ITO北京西单大悦城',
    '体验店_伊稻_门店_ITOITC店',
    '经销_疆远_电商_奕喆科技_可乐惠淘宝代销店',
    '经销_乐绘_家居生活_北京和明兴',
    '经销_伊稻_时尚潮流_深圳市丰瀛盛煌',
    '经销_伊诚_礼品_北京和明兴',
    '联营_乐绘_北京和明兴_ITO重庆万象城店',
    '经销_乐绘_电商_拼多多ITO运动户外亿驰诚专卖店',
    '联营_乐绘_博洛晟_ITO杭州武林银泰店',
    '经销_乐绘_电商_上海韶瑶电子商务_拼多多ITO韶瑶专卖店',
    '联营_乐绘_北京和明兴_ITO重庆龙湖时代天街店',
    '联营_乐绘_北京和明兴_ITO南京新街口金鹰店',
    '联营_乐绘_杭州恒洲_ITO成都IFS',
    '联营_乐绘_北京和明兴_ITO青岛万象城店',
    '经销_乐绘_电商_上海信橙贸易_ITO品牌直售店',
    '直销_有底_电商_有底天猫旗舰店',
    '直销_伊稻_微信小程序ITO商城',
    '体验店_乐绘_门店_ITO上海静安嘉里店',
    '直销_乐绘_电商_抖音ITO官方旗舰店',  # 这个覆盖在了轻熟质享客
    'KA商家_伊诚_大客户',
    '经销_乐绘_电商_天猫ITO乐绘专营店_奕喆科技',
    '经销_乐绘_银行商城_招商银行网上商城_舜冠网络（上海）',
    '经销_乐绘_电商_上海韶瑶电子商务_唯品会ITO旗舰店',
    '联营_乐绘_杨晖_ITO宁波万象城店',
    '联营_乐绘_杭州恒洲_ITO上海浦东嘉里',
    '联营_伊稻_成都王府井百货_ITO成都王府井店',
    '联营_乐绘_杭州恒洲_杭州恒隆',
    '直销_伊稻_电商_视频号ITO旗舰店',
    '直销_有底_电商_有底抖音小店',
    '体验店_乐绘_门店_ITO深圳湾万象城店',
    '经销_伊远_电商_天猫ito伊远专卖店_云礼',
    '经销_伊稻_电商_广发银行网上商城_同齐贸易（上海）',
    '经销_伊诚_礼品_上海云礼',
    '经销_乐绘_福利生活_上海东福网络科技',
    '联营_乐绘_杨晖_ITO温州滨江万象城店',
    '经销_乐绘_电商_上海信橙贸易_ITO品牌直销店',
    '维修_ITO',
    '经销_伊诚_礼品_杭州恒洲',
    '联营_乐绘_杭州晓熙_ITO合肥银泰店',
    '直销_有底_电商_有底小红书旗舰店',
    '换新_ITO',
    '经销_乐绘_电商_抖音ITO箱包旗舰店_奥特传媒',  # 这是达播，暂不处理
    '直销_伊稻_电商_小红书ITO旗舰店',
}


def sync_month(ym, target_wb):
    """读取 YY-M.xlsx 并更新目标文件中的实际销售额"""
    source_file = os.path.join(BASE, '各渠道销售数据源', f'{ym}.xlsx')
    if not os.path.exists(source_file):
        print(f'  [WARN] 文件不存在: {ym}.xlsx，跳过')
        return
    
    df = pd.read_excel(source_file)
    df['日期'] = pd.to_datetime(df['日期'])
    
    year = 2000 + int(ym.split('-')[0])
    month = int(ym.split('-')[1])
    
    # 按店铺+日期汇总实际销售额
    summary = df.groupby(['店铺', df['日期'].dt.date])['实际销售额'].sum().reset_index()
    
    updated_count = 0
    
    for store_name, sheet_name in STORE_TO_SHEET.items():
        ws = target_wb[sheet_name]
        store_data = summary[summary['店铺'] == store_name]
        
        for _, row in store_data.iterrows():
            date_obj = row['日期']
            if isinstance(date_obj, datetime):
                pass
            elif isinstance(date_obj, pd.Timestamp):
                date_obj = date_obj.to_pydatetime()
            else:
                date_obj = datetime.combine(date_obj, datetime.min.time())
            
            actual_val = float(row['实际销售额'])
            
            # 在目标sheet中找到对应日期行
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=1).value
                if cell_val is None:
                    continue
                if hasattr(cell_val, 'strftime'):
                    cell_date = cell_val
                else:
                    try:
                        cell_date = pd.to_datetime(cell_val).to_pydatetime()
                    except:
                        continue
                
                if (cell_date.year == date_obj.year and 
                    cell_date.month == date_obj.month and 
                    cell_date.day == date_obj.day):
                    # 只更新非负的合理值（排除数据错误）
                    if abs(actual_val) > 0:
                        old_val = ws.cell(row=r, column=2).value or 0
                        ws.cell(row=r, column=2).value = actual_val
                        if float(old_val) != actual_val:
                            updated_count += 1
                    break
    
    print(f'  {ym}: 更新了 {updated_count} 个单元格')


def main():
    print('=== 实际销售额同步 ===')
    print()
    
    # 找到所有需要同步的月度数据文件
    src_dir = os.path.join(BASE, '各渠道销售数据源')
    monthly_files = sorted([f for f in os.listdir(src_dir) if f.endswith('.xlsx') and '-' in f and f.startswith('2')])
    
    # 只处理最近修改的文件
    recent_files = []
    for fname in monthly_files:
        fp = os.path.join(src_dir, fname)
        st = os.stat(fp)
        if (time.time() - st.st_mtime) < RECENT_DAYS * 86400:
            recent_files.append(fname)
    
    print(f'找到 {len(monthly_files)} 个数据源文件，其中最近{RECENT_DAYS}天内更新: {len(recent_files)}个')
    if not recent_files:
        print('  没有需要同步的近期文件，跳过')
        return
    
    # 打开目标文件
    target_wb = openpyxl.load_workbook(TARGET_FILE)
    
    for fname in recent_files:
        ym = fname.replace('.xlsx', '')  # e.g., "26-5"
        sync_month(ym, target_wb)
    
    # 保存
    target_wb.save(TARGET_FILE)
    print()
    print('[OK] 同步完成！目标文件已更新')
    print()
    print('下一步：运行 刷新数据.bat 更新看板')


if __name__ == '__main__':
    main()
